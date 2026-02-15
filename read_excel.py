import openpyxl

wb = openpyxl.load_workbook(
    r'I:\我的雲端硬碟\Antigravity\PackListToWMSbyAI\2026 1月攝影方式演練 的複本.xlsx',
    data_only=True
)

with open(r'I:\我的雲端硬碟\Antigravity\PackListToWMSbyAI\excel_structure.txt', 'w', encoding='utf-8') as f:
    f.write(f"Sheets: {wb.sheetnames}\n\n")
    for s in wb.sheetnames:
        ws = wb[s]
        f.write(f"=== Sheet: {s} (Rows={ws.max_row}, Cols={ws.max_column}) ===\n")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)), 1):
            vals = []
            for c in row[:15]:
                v = str(c.value)[:30] if c.value is not None else ""
                vals.append(v)
            f.write(f"  R{i}: {' | '.join(vals)}\n")
        f.write("\n")

print("Done!")
